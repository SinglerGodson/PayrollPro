import json

import pandas
from openpyxl.reader.excel import load_workbook

from src.api.msg_sender import send_msg
from src.api.user import get_by_user_id

user_id_col = 38
config = {
        "教职工": {
            "header_row": 3,
            "name_col": 2, # 姓名列
            "skip_cols": lambda i: i in (1, 3, 4, 38) or i > 28
        },
        "管理人员": {
            "header_row": 5,
            "name_col": 1, # 姓名列
            "skip_cols": lambda i: i == 0 or i > 22
        },
        "工勤人员（校医、教官等）": {
            "header_row": 3,
            "name_col": 1, # 姓名列
            "skip_cols": lambda i:i == 0 or i > 20
        },
        "校长": {
            "header_row": 5,
            "name_col": 1, # 姓名列
            "skip_cols": lambda i: i == 0 or i > 19
        },
        "教职工_高三期绩效": {
            "header_row": 4,
            "name_col": 2, # 姓名列
            "skip_cols": lambda i: i == 1 or i > 9
        },
    }
class ExcelReader:

    def __init__(self, access_token, agent_id, user_input_path):
        self.agent_id = agent_id
        self.access_token = access_token
        self.user_input_path = user_input_path

    def read_write(self):
        print(f'开始加载文件...{self.user_input_path}')
        book = load_workbook(self.user_input_path, data_only=True)
        print('加载文件完毕...')
        for sheet_name in book.sheetnames:
            print(f'处理{sheet_name}')
            if sheet_name not in config:
                print(f'{sheet_name}无效，不处理')
                continue

            sheet = book[sheet_name]
            conf = config[sheet_name]

            skip_cols = conf["skip_cols"]
            header_row = conf["header_row"]

            h1  = sheet[header_row]
            h2 = sheet[header_row + 1]
            msg_col = user_id_col + 1
            send_tag_col = user_id_col + 2
            for row in sheet.iter_rows(min_row=header_row + 2, max_col=send_tag_col + 1):

                if '合计' in str(row[0].value) or '审核' in str(row[0].value):
                    print(f'处理{sheet_name}')
                    break

                row_num = row[0].row
                if row[user_id_col].value is None:
                    print(f'第{row_num}行没有user_id，不处理')
                    continue

                if row[send_tag_col] == "已发送":
                    print(f'第{row_num}行已发送。')
                    continue

                send_msg = row[msg_col].value
                user_id_name = row[user_id_col].value


                user_id_name = str(user_id_name).split('@')
                user_id = user_id_name[1]
                user_name = user_id_name[0]
                if send_msg is None:
                    try:
                        user_info = get_by_user_id(self.access_token, user_id)
                        if user_info["name"].split('-')[0] != user_name:
                            print(f"姓名不匹配。{user_info["name"]} != {user_name}")
                            sheet[f'AO{row_num}'] = f"姓名不匹配。{user_name}"
                            # return "姓名不匹配，未发送"
                    except Exception as e:
                        sheet[f'AO{row_num}'] = str(e)

                    msg_concat = self.msg_concat(h1, h2, row, skip_cols)
                    print(f'填充工资明细信息：{msg_concat}')
                    sheet[f'AN{row_num}'] = msg_concat

                elif row[send_tag_col].value is None:
                    if user_id is not None and not user_id.strip() == "":
                        sheet[f'AO{row_num}'] = self.send_msg(user_id, send_msg)

        print(f'保存文档：{self.user_input_path}')
        book.save(self.user_input_path)



    def msg_concat(self, header1, header2, cells, skip_cols):
        data = {}
        h_1 = None
        for i, cell in enumerate(cells):

            if skip_cols(i):
                continue

            h1 = header1[i].value
            if h1 is None:
                h1 = ""
            if not pandas.isna(h1) and h1 != h_1:
                h_1 = h1

            h_2 = header2[i].value
            value = cell.value
            if pandas.isna(h_2):
                h_2 = h_1

            if not pandas.isna(h_2) and not pandas.isna(value):
                if h_2 == "小计":
                    h_2 = h_1 + h_2

                h_2 = str(h_2).replace("\n", "")
                data[h_2] = value

        return json.dumps(data, ensure_ascii=False).replace('"', "").replace("{", "").replace("}", "")

    def send_msg(self, user_id, msg):
        resp = send_msg(self.access_token, self.agent_id, user_id, msg)
        if resp.status_code != 200:
            return f"发送异常。{resp.text}"
        return "已发送"
